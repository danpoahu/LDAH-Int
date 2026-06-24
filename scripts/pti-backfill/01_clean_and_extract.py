#!/usr/bin/env python3
"""
PTI pink-sheet backfill — step 1 of pipeline (offline, no Firestore).

Reads the source workbook, classifies every activity row, normalizes the
messy fields, and emits two artifacts:

  out/pti_clean_osep.xlsx   Output B — full normalized federal detail for OSEP
                            reporting (every column, all rows), with a
                            Disposition + Flags column, excluded subtotals on
                            their own tab. Detail that does NOT go into the app
                            lives here.

  out/pti_import_source.json  The importable records (NET-NEW / MERGE / DISSEM)
                            with ONLY the fields that map onto existing app
                            schema, plus provenance. Consumed by the Node
                            dry-run / import scripts.

Usage:  python3 01_clean_and_extract.py /path/to/Hawaii_PTI_2025_2030DP.xlsx
"""
import sys, os, json, datetime as dt
from datetime import datetime
import openpyxl

SRC = sys.argv[1] if len(sys.argv) > 1 else \
    "/Users/danielpellegrini/Downloads/Hawaii_PTI_2025_2030DP.xlsx"
HERE = os.path.dirname(os.path.abspath(__file__))
OUT = os.path.join(HERE, "out")
os.makedirs(OUT, exist_ok=True)

# --- grant years that actually hold data -------------------------------------
SHEETS = [
    ("030125_022826", "Year 1 (2025-26)", dt.date(2025, 3, 1), dt.date(2026, 2, 28)),
    ("030126_022827", "Year 2 (2026-27)", dt.date(2026, 3, 1), dt.date(2027, 2, 28)),
]

# Oahu towns / neighbor islands -> the 4 Hawaii counties
COUNTY_MAP = {
    "Honolulu": "Honolulu", "Kaneohe": "Honolulu", "Kailua": "Honolulu", "Leeward": "Honolulu",
    "Hawaii": "Hawaii", "Maui": "Maui", "Molokai": "Maui", "Lanai": "Maui", "Kauai": "Kauai",
}

# header column order (1-indexed -> key)
COLS = [
    "tier_raw", "description", "date", "county_raw", "location", "co_sponsors",
    "dissem_reach", "parents", "prof", "youth", "military", "head_count",
    "age_birth_2", "age_3_5", "age_6_11", "age_12_14", "age_15_18", "age_beyond_hs",
    "eth_afr_am", "eth_asian", "eth_cau", "eth_filip", "eth_hisp", "eth_pt_haw",
    "eth_pac_isl", "eth_other",
    "dis_adhd", "dis_autism", "dis_deaf", "dis_dev_delay", "dis_emo_dist", "dis_intel",
    "dis_gifted", "dis_mult", "dis_no_idea", "dis_ortho", "dis_ohi", "dis_sld",
    "dis_splang", "dis_susp", "dis_tbi", "dis_vis_im", "dis_df_bl",
    "help1_vh", "help1_h", "help1_nh", "help2_vh", "help2_h", "help2_nh",
    "q1_y", "q1_n", "q2_y", "q2_n", "q3_y", "q3_n", "q4_y", "q4_n",
    "q5_y", "q5_n", "q6_y", "q6_n", "q7_y", "q7_n", "q8_y", "q8_n",
]


def norm_tier(t):
    if t is None:
        return (None, "BLANK")
    s = str(t).strip()
    for n in ("1", "2", "3", "4"):
        if s in (f"T{n}", f"T{n}*"):
            return (f"Tier {n}", "ok")
    if s.lower().startswith("dissem"):
        return ("Dissemination", "ok")
    return (None, "JUNK")  # e.g. a date-range typed into the Tier column


def numish(v):
    return v if isinstance(v, (int, float)) else None


def classify(tier_raw, desc, date, hc):
    nt, flag = norm_tier(tier_raw)
    # Non-event roll-up rows: junk tier, OR a date-range typed in the date cell,
    # OR a blank description carrying a big head count.
    if flag == "JUNK" or isinstance(date, str):
        return "EXCLUDE"
    if (desc in (None, "")) and isinstance(hc, (int, float)) and hc and hc > 200:
        return "EXCLUDE"
    d = (desc or "").lower()
    if nt == "Dissemination" or any(k in d for k in
                                    ["dissemination", "facebook", "youtube", "analytics"]):
        return "DISSEM"
    if any(k in d for k in ["connect gen", "ll:", "learning lab", "parent talk", "ptc"]):
        return "MERGE"   # candidate match to an event the live system already has
    return "NETNEW"


def clean_row(raw, grant_label, gy_lo, gy_hi):
    r = {COLS[i]: (raw[i] if i < len(raw) else None) for i in range(len(COLS))}
    flags = []
    assumptions = []

    tier, tflag = norm_tier(r["tier_raw"])
    cls = classify(r["tier_raw"], r["description"], r["date"], r["head_count"])

    # tier noise
    if tflag == "BLANK" and cls != "EXCLUDE":
        tier = "Tier 3"
        assumptions.append("Blank tier inferred as Tier 3 (surrounding TMC group)")
    elif r["tier_raw"] and str(r["tier_raw"]).strip().endswith("*"):
        flags.append(f"Tier asterisk stripped ({r['tier_raw']} -> {tier})")

    # date + grant-year typo handling
    date = r["date"]
    date_iso = None
    grant_year = grant_label
    if isinstance(date, datetime):
        d = date.date()
        if not (gy_lo <= d <= gy_hi):
            # out-of-grant-year. Two known patterns:
            if d.year == gy_hi.year + 0 and d > gy_hi:
                pass
            # TMC Lanai Nov-2026 inside Year-1 sheet -> almost certainly a year typo
            if grant_label.startswith("Year 1") and d.year == 2026 and d.month == 11:
                d = d.replace(year=2025)
                assumptions.append(f"Date corrected {date.date()} -> {d} (year typo)")
            # Oct-2026 LL sessions logged on the Year-1 sheet -> belong to Year 2
            elif grant_label.startswith("Year 1") and d.year == 2026 and d.month >= 3:
                grant_year = "Year 2 (2026-27)"
                assumptions.append(f"Dated {d}; reassigned to Year 2 dataset")
            else:
                flags.append(f"Date {d} outside grant year — review")
        date_iso = d.isoformat()
    elif date not in (None, ""):
        flags.append("Date is not a real date")

    # county
    county = COUNTY_MAP.get(str(r["county_raw"]).strip()) if r["county_raw"] else None
    if r["county_raw"] and not county:
        flags.append(f"Unmapped county '{r['county_raw']}'")
    if not r["county_raw"] and cls != "EXCLUDE":
        county = None  # left "Not recorded"

    r.update(dict(tier=tier, disposition=cls, county=county, date_iso=date_iso,
                  grant_year=grant_year, flags="; ".join(flags),
                  assumptions="; ".join(assumptions)))
    return r


def main():
    wb = openpyxl.load_workbook(SRC, read_only=True, data_only=True)
    all_rows = []
    for sheet, label, lo, hi in SHEETS:
        ws = wb[sheet]
        for raw in ws.iter_rows(min_row=2, values_only=True):
            if not (raw[0] not in (None, "") or raw[1] not in (None, "")
                    or raw[2] not in (None, "")):
                continue
            all_rows.append(clean_row(list(raw), label, lo, hi))

    # ---- Output B: clean OSEP spreadsheet ----
    out_wb = openpyxl.Workbook()
    ws_main = out_wb.active
    ws_main.title = "Activities (clean)"
    ws_excl = out_wb.create_sheet("Excluded subtotals")
    header = (["Grant Year", "Disposition", "Tier", "Date", "County",
               "Description", "Location", "Co-Sponsors", "Dissem Reach",
               "Head Count", "Parents", "Prof", "Youth", "Military"]
              + ["Birth-2", "3-5", "6-11", "12-14", "15-18", "Beyond HS"]
              + ["AfrAm", "Asian", "Cau", "Filip", "Hisp", "PtHaw", "PacIsl", "Other"]
              + ["ADHD", "Autism", "Deaf", "DevDelay", "EmoDist", "Intel", "Gifted",
                 "MultDis", "NoIDEA", "Ortho", "OHI", "SLD", "SpLang", "Susp", "TBI",
                 "VisIm", "DF/BL"]
              + ["Help1-VH", "Help1-H", "Help1-NH", "Help2-VH", "Help2-H", "Help2-NH"]
              + [f"Q{n}-{yn}" for n in range(1, 9) for yn in ("Y", "N")]
              + ["Flags", "Assumptions"])
    for ws in (ws_main, ws_excl):
        ws.append(header)

    demo_keys = (["age_birth_2", "age_3_5", "age_6_11", "age_12_14", "age_15_18",
                  "age_beyond_hs", "eth_afr_am", "eth_asian", "eth_cau", "eth_filip",
                  "eth_hisp", "eth_pt_haw", "eth_pac_isl", "eth_other", "dis_adhd",
                  "dis_autism", "dis_deaf", "dis_dev_delay", "dis_emo_dist", "dis_intel",
                  "dis_gifted", "dis_mult", "dis_no_idea", "dis_ortho", "dis_ohi",
                  "dis_sld", "dis_splang", "dis_susp", "dis_tbi", "dis_vis_im",
                  "dis_df_bl", "help1_vh", "help1_h", "help1_nh", "help2_vh", "help2_h",
                  "help2_nh"]
                 + [f"q{n}_{yn}" for n in range(1, 9) for yn in ("y", "n")])
    for r in all_rows:
        row = ([r["grant_year"], r["disposition"], r["tier"], r["date_iso"],
                r["county"], r["description"], r["location"], r["co_sponsors"],
                numish(r["dissem_reach"]), numish(r["head_count"]),
                numish(r["parents"]), numish(r["prof"]), numish(r["youth"]),
                numish(r["military"])]
               + [numish(r[k]) for k in demo_keys]
               + [r["flags"], r["assumptions"]])
        (ws_excl if r["disposition"] == "EXCLUDE" else ws_main).append(row)
    out_wb.save(os.path.join(OUT, "pti_clean_osep.xlsx"))

    # ---- import-source JSON (only app-mappable fields + provenance) ----
    importable = []
    for i, r in enumerate(all_rows):
        if r["disposition"] == "EXCLUDE":
            continue
        importable.append({
            "rowRef": f"{r['grant_year'][:6].strip()}#{i}",
            "disposition": r["disposition"],          # NETNEW | MERGE | DISSEM
            "grantYear": r["grant_year"],
            "title": r["description"],
            "eventDate": r["date_iso"],
            "location": r["location"] or r["county_raw"] or "",
            "county": r["county"],
            "tierModel": r["tier"],
            # app-mappable counts only:
            "headCount": numish(r["head_count"]),
            "parents": numish(r["parents"]),
            "prof": numish(r["prof"]),
            "youth": numish(r["youth"]),
            "military": numish(r["military"]),
            "dissemReach": numish(r["dissem_reach"]),
            "flags": r["flags"],
            "assumptions": r["assumptions"],
        })
    with open(os.path.join(OUT, "pti_import_source.json"), "w") as f:
        json.dump(importable, f, indent=2)

    # summary to stdout
    from collections import Counter
    c = Counter(x["disposition"] for x in importable)
    print(f"rows total: {len(all_rows)}  importable: {len(importable)}")
    print(f"  NET-NEW {c['NETNEW']}   MERGE {c['MERGE']}   DISSEM {c['DISSEM']}"
          f"   (EXCLUDED {sum(1 for r in all_rows if r['disposition']=='EXCLUDE')})")
    print(f"wrote {OUT}/pti_clean_osep.xlsx")
    print(f"wrote {OUT}/pti_import_source.json")


if __name__ == "__main__":
    main()
